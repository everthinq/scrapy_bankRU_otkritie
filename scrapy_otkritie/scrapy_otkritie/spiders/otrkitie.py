import json
import scrapy
from openpyxl import load_workbook


class otkritie(scrapy.Spider):
    name = 'otkritie'

    def start_requests(self):
        self.cell_value = '2'
        self.workbook = load_workbook('6. PJSC Bank Otkritie Financial Corporation Russia.xlsx')
        self.worksheet = self.workbook[self.workbook.sheetnames[0]]

        for each in range(1, 1000000):
            yield scrapy.Request(
                'https://www.open.ru/api/v2/map/offices/' + str(each),
                callback=self.parse_json
            )

    def parse_json(self, response):
        if (response.text):
            JSON = json.loads(response.text)

            branch_name = JSON['subtitle']
            address = JSON['address']['address']
            lat = JSON['lat']
            lng = JSON['lng']

            print('Writing:', 'id:', JSON['code'], 'name:', branch_name, 'address:', address, 'lat:', lat, 'lng:', lng)

            self.cell_value = str(self.cell_value)
            self.worksheet['B' + self.cell_value] = 'PJSC Bank Otkritie'
            self.worksheet['C' + self.cell_value] = branch_name
            self.worksheet['D' + self.cell_value] = address
            self.worksheet['G' + self.cell_value] = 'Russia'
            self.worksheet['H' + self.cell_value] = 'RU'
            self.worksheet['M' + self.cell_value] = lat
            self.worksheet['N' + self.cell_value] = lng
            self.worksheet['O' + self.cell_value] = 'Address'
            self.worksheet['R' + self.cell_value] = 'Bank website'
            self.cell_value = int(self.cell_value) + 1

            self.workbook.save('6. PJSC Bank Otkritie Financial Corporation Russia.xlsx')
